import streamlit as st
import plotly.graph_objects as go
from datetime import datetime
from utils.forecast_utils import run_prophet_forecast, adjust_forecast
from utils.plot_utils import generate_forecast_plot, generate_trend_plot
from utils.data_utils import export_to_excel
from config.mappings import PN_MODEL_MAPPING
from prophet.diagnostics import cross_validation, performance_metrics
import pandas as pd

def render_analysis():
    """
    Affiche la section "Analyse".
    """
    st.subheader("Analyse des prévisions")
    if st.session_state.pn_data:
        pn_options = sorted(
            [f"{pn} ({PN_MODEL_MAPPING.get(pn, 'Inconnu')})" for pn in st.session_state.pn_data.keys()],
            key=lambda x: (
                PN_MODEL_MAPPING.get(x.split(" (")[0], "Inconnu"),
                x.split(" (")[0]
            )
        )
        selected_pn_display = st.session_state.get('selected_pn', st.selectbox("Sélectionner un PN à analyser", pn_options))
        selected_pn = selected_pn_display.split(" (")[0]
        months = st.slider("Mois à prévoir", 1, 24, 12)
        default_start_date = st.session_state.pn_data[selected_pn]['ds'].max() if not st.session_state.pn_data[selected_pn].empty else datetime(2025, 1, 1)
        forecast_start_date = st.date_input("Date de début des prévisions", value=default_start_date, min_value=datetime(2020, 1, 1), max_value=datetime(2030, 12, 31))

        df = st.session_state.pn_data[selected_pn]
        if df.empty:
            st.error("Les données pour ce PN sont vides. Veuillez charger un fichier valide.")
        else:
            model, forecast = run_prophet_forecast(df, months, forecast_start_date)

            # Nettoyage des tendances : une seule valeur par année (la dernière)
            trends_raw = st.session_state.pn_trend.get(selected_pn, {})
            trends = {}
            for year in sorted(trends_raw.keys()):
                val = trends_raw[year]
                if isinstance(val, dict) and "type" in val and "values" in val:
                    # On prend la dernière valeur pour l'année
                    pct = list(val["values"].values())[-1] if val["values"] else 0.0
                    trends[year] = {"type": "linéaire", "values": {int(year): float(pct)}}
                else:
                    try:
                        pct = float(val)
                    except Exception:
                        pct = 0.0
                    trends[year] = {"type": "linéaire", "values": {int(year): pct}}
            enable_trends = st.session_state.pn_trend_enabled.get(selected_pn, False)
            forecast_adjusted = forecast.copy()
            if enable_trends and trends:
                forecast_adjusted = adjust_forecast(forecast, df, trends, forecast_start_year=forecast_start_date.year)

            df_cv = cross_validation(model, horizon='365 days', initial='730 days', period='180 days')
            mae = performance_metrics(df_cv)['mae'].mean() if not df_cv.empty else None

            forecast_start = pd.to_datetime(forecast_start_date)
            forecast_end_date = (forecast_start + pd.offsets.MonthEnd(months)).normalize()

            if enable_trends and trends:
                min_trend_year = min(int(year) for year in trends.keys())
                max_trend_year = max(int(year) for year in trends.keys())
                trend_start = pd.to_datetime(f"{min(min_trend_year, df['ds'].dt.year.min())}-01-01")
                trend_end = pd.to_datetime(f"{forecast_start.year + (months // 12)}-12-31")
            else:
                trend_start = df['ds'].min()
                trend_end = pd.to_datetime(f"{forecast_start.year + (months // 12)}-12-31")
            all_dates = pd.date_range(start=trend_start, end=trend_end, freq='MS').to_frame(index=False, name='ds')
            trend_forecast = model.predict(all_dates)
            trend_forecast_adjusted = trend_forecast if not (enable_trends and trends) else adjust_forecast(trend_forecast, df, trends, apply_all_trends=True)

            st.markdown(f"### Analyse du PN : **{selected_pn} ({PN_MODEL_MAPPING.get(selected_pn, 'Inconnu')})**")
            st.markdown(f"**Dernière mise à jour** : {st.session_state.pn_last_updated.get(selected_pn)}")
            st.markdown(f"**Utilisation des tendances personnalisées** : {'Activée' if enable_trends else 'Désactivée'}")
            # Affichage clair des tendances personnalisées
            if enable_trends and trends:
                trends_str = ", ".join([
                    f"{year}: {list(val['values'].values())[-1]}%" for year, val in trends.items()
                ])
                st.markdown(f"**Tendances personnalisées** : {trends_str}")
            st.markdown(f"**Début des prévisions** : {forecast_start_date.strftime('%Y-%m-%d')}")
            st.markdown(f"**Fin des prévisions** : {forecast_end_date.strftime('%Y-%m-%d')}")

            st.markdown("#### Indicateurs clés")
            yearly_totals = df.groupby(df['ds'].dt.year)['y'].sum()
            current_year = datetime.now().year
            complete_years = yearly_totals[yearly_totals.index < current_year]
            reference_year = complete_years.index.min() if not complete_years.empty else None
            last_complete_year = complete_years.index.max() if not complete_years.empty else None
            total_growth = 0
            if reference_year is not None and last_complete_year is not None and reference_year != last_complete_year:
                reference_total = complete_years.get(reference_year, 0)
                last_complete_total = complete_years.get(last_complete_year, 0)
                total_growth = ((last_complete_total - reference_total) / reference_total * 100) if reference_total != 0 else 0

            previous_year = forecast_start_date.year - 1
            total_previous_year = yearly_totals.get(previous_year, 0)
            total_forecast_period = forecast_adjusted['yhat'].sum() if not forecast_adjusted.empty else 0
            monthly_avg_forecast_period = total_forecast_period / months if total_forecast_period != 0 else 0

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric(
                "Croissance totale",
                f"{total_growth:.1f}%",
                delta=f"{total_growth:.1f}%",
                delta_color="normal" if total_growth >= 0 else "inverse",
                help=f"Pourcentage de croissance entre {reference_year} et {last_complete_year}" if reference_year and last_complete_year else "Pas assez de données"
            )
            col2.metric(
                f"Total {previous_year}",
                f"{total_previous_year:.0f}",
                help=f"Quantité totale pour l’année {previous_year}"
            )
            col3.metric(
                f"Total prévu ({forecast_start.strftime('%Y-%m')} à {forecast_end_date.strftime('%Y-%m')})",
                f"{total_forecast_period:.0f}",
                help=f"Quantité totale prévue pour la période de {forecast_start_date.strftime('%Y-%m-%d')} à {forecast_end_date.strftime('%Y-%m-%d')}"
            )
            col4.metric(
                f"Moyenne mensuelle ({forecast_start.strftime('%Y-%m')} à {forecast_end_date.strftime('%Y-%m')})",
                f"{monthly_avg_forecast_period:.0f}",
                help=f"Moyenne mensuelle prévue pour la période de {forecast_start_date.strftime('%Y-%m-%d')} à {forecast_end_date.strftime('%Y-%m-%d')}"
            )
            col5.metric(
                "Fiabilité (MAE)",
                f"{mae:.1f}" if mae is not None else "N/A",
                help="Erreur moyenne absolue des prévisions basée sur la validation croisée"
            )

            yearly_totals = df[df['ds'].dt.year < current_year].groupby(df['ds'].dt.year)['y'].sum().reset_index()
            yearly_totals.columns = ['year', 'total']
            growth_text = "Pas assez de données historiques pour calculer la croissance entre 2023 et 2024."
            if len(yearly_totals) >= 2:
                last_year = yearly_totals['year'].iloc[-1]
                second_last_year = yearly_totals['year'].iloc[-2]
                if last_year == 2024 and second_last_year == 2023:
                    last_total = yearly_totals['total'].iloc[-1]
                    second_last_total = yearly_totals['total'].iloc[-2]
                    growth_last_two_years = ((last_total - second_last_total) / second_last_total * 100) if second_last_total != 0 else 0
                    growth_text = f"Croissance 2023 à 2024: {growth_last_two_years:.1f}%"

            fig_growth = go.Figure(data=[
                go.Bar(
                    x=yearly_totals['year'],
                    y=yearly_totals['total'],
                    marker_color='#4A90E2',
                    text=[f"{total:.0f}" for total in yearly_totals['total']],
                    textposition='auto',
                    hovertemplate='Année: %{x}<br>Quantité: %{y:.0f}<extra></extra>'
                )
            ])
            fig_growth.update_layout(
                title=f'Quantités annuelles historiques ({growth_text})',
                xaxis_title='Année',
                yaxis_title='Quantité totale',
                height=400,
                showlegend=False,
                xaxis=dict(tickmode='linear', tick0=yearly_totals['year'].min() if not yearly_totals.empty else 2020, dtick=1),
                plot_bgcolor='#F5F7FA',
                paper_bgcolor='#FFFFFF',
                font_color='#003087'
            )
            st.plotly_chart(fig_growth, use_container_width=True)

            # Prévisions initiale et impactée par trend perso
            forecast_initial = forecast.copy()
            forecast_impact = forecast_adjusted.copy() if (enable_trends and trends) else forecast.copy()
            # Affichage du graphique principal : deux scénarios si trend activée
            st.markdown("#### Prévision de la demande")
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=df['ds'], y=df['y'], mode='lines+markers', name='Données réelles',
                line=dict(color='#003087'), marker=dict(size=10),
                hovertemplate='Date: %{x|%Y-%m}<br>Quantité: %{y:.0f}<extra></extra>'
            ))
            fig.add_trace(go.Scatter(
                x=forecast_initial['ds'], y=forecast_initial['yhat'], name='Prévision initiale',
                line=dict(dash='dash', color='orange'),
                hovertemplate='Date: %{x|%Y-%m}<br>Prévision: %{y:.0f}<extra></extra>'
            ))
            if enable_trends and trends:
                fig.add_trace(go.Scatter(
                    x=forecast_impact['ds'], y=forecast_impact['yhat'], name='Prévision trend perso',
                    line=dict(color='firebrick'),
                    hovertemplate='Date: %{x|%Y-%m}<br>Prévision trend perso: %{y:.0f}<extra></extra>'
                ))
            fig.add_trace(go.Scatter(
                x=forecast_initial['ds'], y=forecast_initial['yhat_upper'], fill=None, mode='lines', line_color='rgba(0,0,0,0)', showlegend=False))
            fig.add_trace(go.Scatter(
                x=forecast_initial['ds'], y=forecast_initial['yhat_lower'], fill='tonexty', fillcolor='rgba(0,176,246,0.2)', mode='lines', line_color='rgba(0,0,0,0)', name='Intervalle de confiance', hovertemplate='Date: %{x|%Y-%m}<br>Intervalle: %{y:.0f}<extra></extra>'))
            if not df['ds'].empty:
                last_historical_date = df['ds'].max()
                fig.add_vline(x=last_historical_date.timestamp() * 1000, line=dict(color='#CE1126', dash='dash'), annotation_text="Début des données historiques", annotation_position="top")
            fig.update_layout(
                title=f'Prévisions pour {selected_pn} ({PN_MODEL_MAPPING.get(selected_pn, "Inconnu")})',
                xaxis_title='Date',
                yaxis_title='Quantité',
                height=600,
                showlegend=True,
                margin=dict(l=50, r=50, t=50, b=50),
                xaxis=dict(type='date'),
                plot_bgcolor='#F5F7FA',
                paper_bgcolor='#FFFFFF',
                font_color='#003087'
            )
            st.plotly_chart(fig, use_container_width=True)
            # Affichage du graph de trend seule
            st.markdown("#### Visualisation de la trend")
            trend_forecast = model.predict(all_dates)
            trend_forecast_adjusted = trend_forecast if not (enable_trends and trends) else adjust_forecast(trend_forecast, df, trends, apply_all_trends=True)
            fig_trend = go.Figure()
            fig_trend.add_trace(go.Scatter(x=trend_forecast['ds'], y=trend_forecast['trend'], name='Trend initiale', line=dict(dash='dash', color='orange')))
            if enable_trends and trends:
                fig_trend.add_trace(go.Scatter(x=trend_forecast_adjusted['ds'], y=trend_forecast_adjusted['trend'], name='Trend perso', line=dict(color='firebrick')))
            fig_trend.update_layout(title='Trend initiale et trend perso', xaxis_title='Date', yaxis_title='Tendance', height=400, showlegend=True, plot_bgcolor='#F5F7FA', paper_bgcolor='#FFFFFF', font_color='#003087')
            st.plotly_chart(fig_trend, use_container_width=True)

            # Préparation du tableau Excel prévision
            forecast_out = forecast_adjusted[['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy()
            forecast_out['Mois'] = forecast_out['ds'].dt.strftime('%B')
            forecast_out = forecast_out[['Mois', 'yhat', 'yhat_lower', 'yhat_upper']]
            forecast_out.columns = ['Mois', 'Prévision', 'Valeur basse', 'Valeur haute']
            forecast_out['Prévision'] = forecast_out['Prévision'].round(0).astype(int)
            forecast_out['Valeur basse'] = forecast_out['Valeur basse'].round(0).astype(int)
            forecast_out['Valeur haute'] = forecast_out['Valeur haute'].round(0).astype(int)

            # Génération Excel customisée
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            output = BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Prévisions'

            # Titre fusionné
            titre = f"PN : {selected_pn} | Période de prévision : {forecast_start.strftime('%B %Y')} à {forecast_end_date.strftime('%B %Y')} | Date de génération : {datetime.now().strftime('%d %B %Y')}"
            worksheet['A1'] = titre
            worksheet.merge_cells('A1:D1')
            worksheet['A1'].font = Font(name='Helvetica', size=12, bold=True, color='003087')
            worksheet['A1'].alignment = Alignment(horizontal='left')
            worksheet['A1'].fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
            worksheet['A1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # En-têtes
            headers = ['Mois', 'Prévision', 'Valeur basse', 'Valeur haute']
            for col_num, header in enumerate(headers, 1):
                worksheet.cell(row=2, column=col_num).value = header
                worksheet.cell(row=2, column=col_num).fill = PatternFill(start_color='4A4A4A', end_color='4A4A4A', fill_type='solid')
                worksheet.cell(row=2, column=col_num).font = Font(color='FFFFFF', bold=True)
                worksheet.cell(row=2, column=col_num).alignment = Alignment(horizontal='center')
                worksheet.cell(row=2, column=col_num).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Données
            for row_num, row_data in enumerate(forecast_out.values, 3):
                for col_num, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col_num).value = value
                    worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
                    worksheet.cell(row=row_num, column=col_num).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    worksheet.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

            # Largeur colonnes
            for col_num in range(1, 5):
                max_length = 0
                column_letter = get_column_letter(col_num)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_num)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output)
            excel_data = output.getvalue()

            st.download_button(
                label="Télécharger l’analyse",
                data=excel_data,
                file_name=f"analyse_{selected_pn}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_analysis",
                use_container_width=True
            )
    else:
        st.info("Ajoutez un PN pour démarrer l’analyse.")

    st.markdown("---")